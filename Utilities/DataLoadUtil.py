import openpyxl


class DataLoadUtil:

    @staticmethod
    def data_load(sheet_name):
        workbook = openpyxl.load_workbook("../TestData/testData.xlsx")
        sheet = workbook[sheet_name]

        main_data = []
        for i in range(2, sheet.max_row + 1):
            data = []
            for j in range(1, sheet.max_column + 1):
                data.append(sheet.cell(row=i, column=j).value)
            main_data.insert(i - 2, data)

        return main_data
